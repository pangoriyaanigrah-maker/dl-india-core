"""Excel/CSV -> plain dicts, plus the validation that goes with it.

Two rules here are easy to get wrong and both were bugs once:

  * the header row is FOUND, not assumed to be row 1 -- the DL template
    has it part-way down the sheet
  * percent-vs-fraction for weights is decided ONCE PER COLUMN from the
    column total, never per cell. Judging each cell against a 1.5 cutoff
    reads a legitimate 1.2% position as 120%.
"""
from __future__ import annotations

import csv
import datetime as dt
import io
import re

import openpyxl

VALID_SIDES = {"Buy", "Sell"}
VALID_CF_TYPES = {"Contribution", "Withdrawal"}
HEADER_KEYS = {"ticker", "nseticker", "symbol", "company"}
CASHFLOW_HEADER_KEYS = {"date", "type"}


class ImportError_(ValueError):
    """Message is meant to be read by the person doing the import."""


# ------------------------------------------------------------ scalars
def norm(cell) -> str:
    return re.sub(r"[^a-z0-9]", "", str(cell or "").lower())


def num(v):
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v)
    for ch in ",₹% \t":
        s = s.replace(ch, "")
    try:
        return float(s)
    except ValueError:
        return None


def ticker(raw) -> str:
    if raw is None or not str(raw).strip():
        raise ImportError_("Ticker must not be empty.")
    tk = str(raw).strip().upper()
    # A ticker never contains a space; a sector or company name usually
    # does. Seeing one means the header matched the wrong column, and
    # importing would quietly replace the book with nonsense.
    if re.search(r"\s", tk):
        raise ImportError_(
            f"Column mapping looks wrong — this was read as a ticker: {tk!r}. "
            f"Check the sheet has a 'Ticker' header above the data."
        )
    return tk


def trade_date(raw) -> dt.date:
    if raw is None:
        raise ImportError_("Trade date is required.")
    if isinstance(raw, dt.datetime):
        return raw.date()
    if isinstance(raw, dt.date):
        return raw
    try:
        return dt.datetime.fromisoformat(str(raw)).date()
    except ValueError:
        raise ImportError_(f"Could not parse trade date: {raw!r}")


def side(raw) -> str:
    s = str(raw or "").strip().capitalize()
    if s not in VALID_SIDES:
        raise ImportError_(f"Side must be Buy or Sell, got {raw!r}.")
    return s


def cf_type(raw) -> str:
    s = str(raw or "").strip().capitalize()
    if s not in VALID_CF_TYPES:
        raise ImportError_(f"Type must be Contribution or Withdrawal, got {raw!r}.")
    return s


def positive(raw, label) -> float:
    v = num(raw)
    if v is None:
        raise ImportError_(f"{label} must be numeric, got {raw!r}.")
    if v <= 0:
        raise ImportError_(f"{label} must be > 0, got {v}.")
    return v


# ------------------------------------------------------------ columns
def column(header, *names) -> int:
    normalized = [norm(c) for c in header]
    for n in names:
        if n in normalized:
            return normalized.index(n)
    return -1


def holdings_columns(header) -> dict:
    col = {
        "tk": column(header, "ticker", "nseticker", "symbol"),
        "co": column(header, "company", "name"),
        "sector": column(header, "sector"),
        "industry": column(header, "industry", "subsector", "theme"),
        "analyst": column(header, "analyst"),
        "tp": column(header, "target", "targetprice", "tp"),
        "qty": column(header, "qty", "quantity", "shares", "sharesheld"),
        "wt": column(header, "weight", "wt"),
        "fullwt": column(header, "fullweight", "fullwt", "targetweight"),
        "addlvl": column(header, "addlevel", "addlvl", "addprice"),
        "conv": column(header, "conviction", "conv"),
        "thesis": column(header, "thesis", "comments", "notes"),
        "strategy": column(header, "strategy", "plan"),
        "yfsymbol": column(header, "yfsymbol", "yahoosymbol", "yahooticker"),
    }
    if col["tk"] < 0 and col["co"] < 0:
        raise ImportError_("Holdings file needs a Ticker or Company column.")
    if col["tp"] < 0:
        raise ImportError_("Holdings file needs a Target column.")
    return col


def trades_columns(header) -> dict:
    col = {
        "date": column(header, "date"),
        "tk": column(header, "ticker", "symbol"),
        "side": column(header, "side"),
        "qty": column(header, "qty", "quantity", "shares"),
        "price": column(header, "price"),
        "costs": column(header, "costs", "fees", "charges"),
    }
    if any(col[n] < 0 for n in ("tk", "side", "qty", "price")):
        raise ImportError_("Trades file needs Ticker, Side, Qty and Price columns.")
    return col


def cashflow_columns(header) -> dict:
    col = {
        "date": column(header, "date"),
        "type": column(header, "type"),
        "amount": column(header, "amount", "value"),
        "note": column(header, "note", "notes", "comment", "comments"),
    }
    if any(col[n] < 0 for n in ("date", "type", "amount")):
        raise ImportError_("Cashflows file needs Date, Type and Amount columns.")
    return col


# -------------------------------------------------------------- sheets
def rows_from(data: bytes, filename: str, sheet: str, header_keys=HEADER_KEYS) -> list[list]:
    """-> rows starting AT the header row."""
    if filename.lower().endswith(".csv"):
        rows = list(csv.reader(io.StringIO(data.decode("utf-8-sig", errors="replace"))))
    else:
        try:
            wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
        except Exception as e:
            raise ImportError_(f"Could not read this workbook ({e}). Expected .xlsx or .csv.")
        if sheet not in wb.sheetnames:
            raise ImportError_(
                f"No '{sheet}' sheet in this workbook. Found: {', '.join(wb.sheetnames)}."
            )
        rows = [list(r) for r in wb[sheet].iter_rows(values_only=True)]

    if not rows:
        raise ImportError_("Empty file.")
    for i, r in enumerate(rows):
        if any(norm(c) in header_keys for c in (r or [])):
            return rows[i:]
    raise ImportError_(f"No header row found — need a {' or '.join(sorted(header_keys))} column.")


def _scale(body, col) -> float:
    """Fractions (0.025) or percents (2.5)? Decided once, from the column
    total: book weights sum to ~1 as fractions, ~100 as percents."""
    if col < 0:
        return 1.0
    vals = [v for v in (num(r[col] if col < len(r) else None) for r in body) if v is not None]
    return 100.0 if sum(vals) > 1.5 else 1.0


def parse_holdings(data: bytes, filename: str) -> tuple[list[dict], list[dict]]:
    """-> (holdings, row errors). A bad row is reported with its row number
    and skipped; it never aborts the whole import."""
    rows = rows_from(data, filename, "Portfolio")
    col = holdings_columns(rows[0])
    g = lambda r, c: (r[c] if 0 <= c < len(r) else None)  # noqa: E731

    body = [r for r in rows[1:] if g(r, col["tk"]) or g(r, col["co"])]
    w_div = _scale(body, col["wt"])
    f_div = _scale(body, col["fullwt"]) if col["fullwt"] > -1 else w_div

    out, errors = [], []
    for i, r in enumerate(body, start=1):
        try:
            tk = ticker(g(r, col["tk"]) or g(r, col["co"]))
        except ImportError_ as e:
            errors.append({"row": i, "message": str(e)})
            continue
        wt = num(g(r, col["wt"]))
        wt = (wt / w_div) if wt is not None else 0.0
        full = num(g(r, col["fullwt"]))
        full = (full / f_div) if full is not None else wt
        yf = g(r, col["yfsymbol"])
        out.append({
            "tk": tk,
            "co": (str(g(r, col["co"]) or tk).strip() or None),
            "sector": g(r, col["sector"]), "industry": g(r, col["industry"]),
            "analyst": g(r, col["analyst"]), "tp": num(g(r, col["tp"])),
            "qty": num(g(r, col["qty"])),
            "wt": wt, "fullWt": full, "addLvl": num(g(r, col["addlvl"])),
            "yfSymbol": str(yf).strip() if yf else None,
            "conv": g(r, col["conv"]),
            "thesis": str(g(r, col["thesis"])) if g(r, col["thesis"]) else None,
            "strategy": str(g(r, col["strategy"])) if g(r, col["strategy"]) else None,
        })
    if not out:
        raise ImportError_("No valid holdings rows found.")
    return out, errors


def parse_trades(data: bytes, filename: str) -> tuple[list[dict], list[dict]]:
    rows = rows_from(data, filename, "Trades")
    col = trades_columns(rows[0])
    g = lambda r, c: (r[c] if 0 <= c < len(r) else None)  # noqa: E731

    out, errors = [], []
    body = [r for r in rows[1:] if g(r, col["tk"])]
    for i, r in enumerate(body, start=1):
        try:
            out.append({
                "tk": ticker(g(r, col["tk"])),
                "date": trade_date(g(r, col["date"])).isoformat(),
                "side": side(g(r, col["side"])),
                "qty": positive(g(r, col["qty"]), "Quantity"),
                "price": positive(g(r, col["price"]), "Price"),
                "costs": num(g(r, col["costs"])) or 0.0,
            })
        except ImportError_ as e:
            errors.append({"row": i, "message": str(e)})
    if not out:
        raise ImportError_("No valid trade rows found.")
    return out, errors


def parse_cashflows(data: bytes, filename: str) -> tuple[list[dict], list[dict]]:
    """Money moving in or out of the book from outside -- the Principal
    funding it or pulling money out. Never a trade: buying/selling a stock
    moves value between cash and a position, it doesn't change the total."""
    rows = rows_from(data, filename, "Cashflows", header_keys=CASHFLOW_HEADER_KEYS)
    col = cashflow_columns(rows[0])
    g = lambda r, c: (r[c] if 0 <= c < len(r) else None)  # noqa: E731

    out, errors = [], []
    body = [r for r in rows[1:] if g(r, col["date"]) or g(r, col["type"])]
    for i, r in enumerate(body, start=1):
        try:
            out.append({
                "date": trade_date(g(r, col["date"])).isoformat(),
                "type": cf_type(g(r, col["type"])),
                "amount": positive(g(r, col["amount"]), "Amount"),
                "note": str(g(r, col["note"])) if g(r, col["note"]) else None,
            })
        except ImportError_ as e:
            errors.append({"row": i, "message": str(e)})
    if not out:
        raise ImportError_("No valid cashflow rows found.")
    return out, errors


def merge_trades(existing: list[dict], incoming: list[dict]) -> tuple[list[dict], int, int]:
    """Trades are a ledger, not a snapshot: a new file ADDS to what is
    already recorded. Deduped on (date, ticker, side, qty, price) so
    re-uploading the same file changes nothing.

    -> (merged, added, duplicates)
    """
    key = lambda t: (t["date"], t["tk"], t["side"], t["qty"], t["price"])  # noqa: E731
    seen = {key(t) for t in existing}
    added = [t for t in incoming if key(t) not in seen]
    return existing + added, len(added), len(incoming) - len(added)


def merge_cashflows(existing: list[dict], incoming: list[dict]) -> tuple[list[dict], int, int]:
    """Same ledger semantics as merge_trades: a new file adds to what's
    already recorded, deduped on (date, type, amount) so a re-upload of
    the same file is a no-op."""
    key = lambda c: (c["date"], c["type"], c["amount"])  # noqa: E731
    seen = {key(c) for c in existing}
    added = [c for c in incoming if key(c) not in seen]
    return existing + added, len(added), len(incoming) - len(added)


# --------------------------------------------------------------- export
# The reverse direction: current book -> a downloadable .xlsx in the same
# shape holdings_columns()/trades_columns() read. Weight/FullWeight are
# written back out as percent numbers (2.5, not 0.025) to match the file
# format the app already expects on re-upload.
def to_holdings_xlsx(holdings: list[dict]):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Portfolio"
    ws.append(["Ticker", "Company", "Sector", "Industry", "Analyst", "Qty", "Weight",
               "FullWeight", "Target", "AddLevel", "Conviction", "Thesis", "Strategy", "YfSymbol"])
    for h in holdings:
        ws.append([h.get("tk"), h.get("co"), h.get("sector"), h.get("industry"), h.get("analyst"),
                   h.get("qty"), (h.get("wt") or 0) * 100, (h.get("fullWt") or 0) * 100,
                   h.get("tp"), h.get("addLvl"), h.get("conv"), h.get("thesis"),
                   h.get("strategy"), h.get("yfSymbol")])
    return wb


def to_trades_xlsx(trades: list[dict]):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Trades"
    ws.append(["Date", "Ticker", "Side", "Qty", "Price", "Costs"])
    for t in trades:
        ws.append([t.get("date"), t.get("tk"), t.get("side"), t.get("qty"),
                   t.get("price"), t.get("costs")])
    return wb


def to_cashflows_xlsx(cashflows: list[dict]):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cashflows"
    ws.append(["Date", "Type", "Amount", "Note"])
    for c in cashflows:
        ws.append([c.get("date"), c.get("type"), c.get("amount"), c.get("note")])
    return wb
