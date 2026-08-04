"""End-to-end against the local storage backend: import real files, check
the maths, check the stored JSON, check the failure paths.

DRIVE_LOCAL_DIR is set before importing the app so `drive.connect()` uses
a temp folder rather than reaching for Google credentials.
"""
import datetime as dt
import io
import json
import os
import tempfile
import urllib.error

import openpyxl
import pytest

TMP = tempfile.mkdtemp(prefix="dl_test_")
os.environ["DRIVE_LOCAL_DIR"] = TMP

from fastapi.testclient import TestClient   # noqa: E402
import calculator                            # noqa: E402
import drive                                 # noqa: E402
import parser                                # noqa: E402
from main import app                         # noqa: E402

HOLDINGS = [
    ["ALPHACHEM", "Alpha Chemicals", "Materials", "Specialty", "Dev", 8.0, 10.0, 620, 480, "High", "t", "s"],
    ["BETAFIN",  "Beta Finserv",    "Financials", "NBFC",     "Dev", 7.0, 0.0,  340, None, "Low",  "t", "s"],
]
H_HDR = ["Ticker", "Company", "Sector", "Industry", "Analyst", "Weight", "FullWeight",
         "Target", "AddLevel", "Conviction", "Thesis", "Strategy"]
T_HDR = ["Date", "Ticker", "Side", "Qty", "Price", "Costs"]
CF_HDR = ["Date", "Type", "Amount", "Note"]
TRADES = [
    [dt.date(2026, 1, 15), "ALPHACHEM", "Buy", 400, 505.0, 202.0],
    [dt.date(2026, 2, 3), "BETAFIN", "Buy", 900, 298.0, 268.2],
]
FEED = {
    "asof": "29 Jul 2026, 18:00 IST", "index": "Test Index",
    "bench_sect": {"Materials": 0.2, "Financials": 0.3}, "bench_size": {"Small": 1.0},
    "errors": [],
    "signals": {
        "ALPHACHEM": {"cmp": 542.0, "mcap": 3400, "lo": 470.0, "hi": 615.0, "val": 0.3,
                       "momo": 0.6, "qual": 0.1, "r1m": 0.04, "worstM": -0.18, "mdd": -0.3,
                       "dvol": 0.02, "beta": 0.9, "adv": 500.0},
        "BETAFIN": {"cmp": 305.0, "mcap": 2100, "lo": 260.0, "hi": 355.0, "val": 0.8,
                     "momo": -0.1, "qual": 0.4, "r1m": -0.02, "worstM": -0.22, "mdd": -0.4,
                     "dvol": 0.03, "beta": 1.1, "adv": 5000.0},
    },
}


def xlsx(sheet, header, rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet
    ws.append(header)
    for r in rows:
        ws.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@pytest.fixture()
def client(monkeypatch):
    for f in os.listdir(os.path.join(TMP, "Portfolio")) if os.path.isdir(os.path.join(TMP, "Portfolio")) else []:
        os.remove(os.path.join(TMP, "Portfolio", f))
    # Every holdings/trades import now fires a real Yahoo Finance call
    # (_quick_refresh_prices) -- stubbed here so the whole suite doesn't
    # pay real network latency (or flakiness) on every single import test.
    # A dedicated test below overrides this to exercise the real merge.
    import build_signals
    monkeypatch.setattr(build_signals, "quick_prices", lambda holdings: ({}, []))
    with TestClient(app) as c:          # startup event runs drive.connect()
        drive.write_json(drive.SIGNALS_JSON, FEED)
        yield c


def up(c, kind, data, name="f.xlsx"):
    return c.post(f"/api/import/{kind}", files={"file": (name, data, "application/octet-stream")})


# ------------------------------------------------------------ storage
def test_startup_creates_the_portfolio_folder_and_every_json_file(client):
    names = drive.store().list_names()
    for f in drive.JSON_FILES:
        assert f in names, f"{f} missing from {names}"


def test_health_reports_storage(client):
    r = client.get("/health").json()
    assert r["status"] == "ok" and r["storage"] == "connected"


def test_empty_install_renders_rather_than_erroring(client):
    for path in ("/api/dashboard", "/api/portfolio", "/api/risk", "/api/performance",
                  "/api/exposure", "/api/positions", "/api/signals"):
        assert client.get(path).status_code == 200, path
    assert client.get("/api/dashboard").json()["holdingsCount"] == 0


# ------------------------------------------------------------- import
def test_holdings_import_stores_xlsx_and_recalculates(client):
    r = up(client, "holdings", xlsx("Portfolio", H_HDR, HOLDINGS))
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["holdings"] == 2
    assert body["cash"] == pytest.approx(1 - 0.15)      # 1 - (0.08 + 0.07)

    assert drive.store().exists(drive.HOLDINGS_XLSX)     # original kept
    book = drive.read_json(drive.PORTFOLIO_JSON)
    assert [h["tk"] for h in book["holdings"]] == ["ALPHACHEM", "BETAFIN"]
    d = client.get("/api/dashboard").json()
    assert d["holdingsCount"] == 2 and d["cashWeight"] == pytest.approx(0.85)


def test_percent_vs_fraction_decided_per_column_not_per_cell(client):
    """A 1.2% position must not be read as 120%."""
    up(client, "holdings", xlsx("Portfolio", ["Ticker", "Company", "Weight", "Target"],
                                 [["AAA", "A", 1.2, 620], ["BBB", "B", 98.8, 340]]))
    pos = {p["tk"]: p for p in client.get("/api/positions").json()["positions"]}
    assert pos["AAA"]["wt"] == pytest.approx(0.012)


def test_trades_append_and_dedupe(client):
    up(client, "holdings", xlsx("Portfolio", H_HDR, HOLDINGS))
    assert up(client, "trades", xlsx("Trades", T_HDR, TRADES)).json()["added"] == 2

    again = up(client, "trades", xlsx("Trades", T_HDR, TRADES + [
        [dt.date(2026, 3, 1), "ALPHACHEM", "Buy", 50, 520.0, 26.0]])).json()
    assert again["added"] == 1 and again["duplicates"] == 2 and again["total"] == 3


def test_trade_for_unknown_ticker_creates_a_placeholder(client):
    up(client, "holdings", xlsx("Portfolio", H_HDR, HOLDINGS))
    body = up(client, "trades", xlsx("Trades", T_HDR,
              [[dt.date(2026, 4, 1), "ZETAENER", "Buy", 10, 640.0, 6.4]])).json()
    assert body["placeholderHoldings"] == ["ZETAENER"]
    z = next(p for p in client.get("/api/positions").json()["positions"] if p["tk"] == "ZETAENER")
    assert z["needsMetadata"] is True


def test_fully_exited_unknown_ticker_gets_no_placeholder(client):
    """A ticker with no holdings row that nets to zero across the trades
    that mention it (bought and fully sold, never held per the holdings
    file) has nothing to manage -- it must not clutter Positions, but its
    P&L stays visible on Performance since that reads the trade ledger
    directly, not the holdings list."""
    up(client, "holdings", xlsx("Portfolio", H_HDR, HOLDINGS))
    body = up(client, "trades", xlsx("Trades", T_HDR, [
        [dt.date(2026, 5, 20), "BETAOLD", "Buy", 300, 300.0, 20.0],
        [dt.date(2026, 5, 28), "BETAOLD", "Sell", 300, 320.0, 20.0],
    ])).json()
    assert body["placeholderHoldings"] == []
    assert not any(p["tk"] == "BETAOLD" for p in client.get("/api/positions").json()["positions"])
    row = next(r for r in client.get("/api/performance").json()["positions"] if r["tk"] == "BETAOLD")
    assert row["realized"] == pytest.approx(6000) and row["net"] == 0
    # its P&L should be clearly labelled, not lumped into a vague bucket
    assert row["sector"] == calculator.EXITED_LABEL
    assert row["bucket"] == calculator.EXITED_LABEL
    port = client.get("/api/portfolio").json()
    assert any(b["key"] == calculator.EXITED_LABEL and b["pl"] == pytest.approx(5960)
               for b in port["attribution"]["bySector"])
    assert any(b["key"] == calculator.EXITED_LABEL and b["pl"] == pytest.approx(5960)
               for b in port["attribution"]["bySize"])


def test_export_round_trips_back_through_the_importer(client):
    """The downloaded .xlsx is not just a dump -- it has to be the exact
    shape the importer itself reads, so 'edit and re-upload' actually
    works. Round-trip holdings and trades through export -> import and the
    book must come out unchanged."""
    up(client, "holdings", xlsx("Portfolio", H_HDR, HOLDINGS))
    up(client, "trades", xlsx("Trades", T_HDR, TRADES))
    before = client.get("/api/dashboard").json()

    h_xlsx = client.get("/api/export/holdings").content
    t_xlsx = client.get("/api/export/trades").content

    up(client, "holdings", h_xlsx)
    up(client, "trades", t_xlsx)
    after = client.get("/api/dashboard").json()

    assert after["holdingsCount"] == before["holdingsCount"] == 2
    assert after["tradesCount"] == before["tradesCount"]           # trades dedupe, not double
    assert after["unrealized"] == pytest.approx(before["unrealized"])
    assert after["realized"] == pytest.approx(before["realized"])
    pos = {p["tk"]: p for p in client.get("/api/positions").json()["positions"]}
    assert pos["ALPHACHEM"]["wt"] == pytest.approx(0.08)


def test_nav_from_real_cashflows_and_trade_cash(client):
    """Rs 5,00,000 contributed. Trade cash out: ALPHACHEM 400@505+202 costs
    = 202,202; BETAFIN 900@298+268.2 = 268,468.2. Cash balance = 500,000 -
    470,670.2 = 29,329.8. Holdings value at live CMP: 400*542 + 900*305 =
    491,300. NAV = 520,629.8 -- and must equal contributed + unrealized -
    fees exactly, an independent cross-check via the P&L identity."""
    up(client, "holdings", xlsx("Portfolio", H_HDR, HOLDINGS))
    up(client, "trades", xlsx("Trades", T_HDR, TRADES))
    r = up(client, "cashflows", xlsx("Cashflows", CF_HDR,
           [[dt.date(2026, 1, 1), "Contribution", 500000, "Seed capital"]]))
    assert r.json() == {"file": "f.xlsx", "added": 1, "duplicates": 0, "total": 1, "errors": []}

    nav = client.get("/api/nav").json()
    assert nav["contributed"] == 500000 and nav["withdrawn"] == 0
    assert nav["tradeCash"] == pytest.approx(-470670.2)
    assert nav["cashBalance"] == pytest.approx(29329.8)
    assert nav["holdingsValue"] == pytest.approx(491300.0)
    assert nav["nav"] == pytest.approx(520629.8)

    d = client.get("/api/dashboard").json()
    assert nav["nav"] == pytest.approx(500000 + d["unrealized"] + d["realized"] - d["costs"])

    # duplicate upload must not double-count the contribution
    r2 = up(client, "cashflows", xlsx("Cashflows", CF_HDR,
            [[dt.date(2026, 1, 1), "Contribution", 500000, "Seed capital"]]))
    assert r2.json()["added"] == 0 and r2.json()["duplicates"] == 1
    assert client.get("/api/nav").json()["contributed"] == 500000

    # a withdrawal must reduce cash balance and NAV by exactly its amount
    up(client, "cashflows", xlsx("Cashflows", CF_HDR,
       [[dt.date(2026, 3, 1), "Withdrawal", 20000, "Partial pull"]]))
    nav2 = client.get("/api/nav").json()
    assert nav2["withdrawn"] == 20000
    assert nav2["nav"] == pytest.approx(nav["nav"] - 20000)


def test_dashboard_equity_cash_weight_uses_real_nav_not_the_typed_plug(client):
    """Real bug this fixes: the Overview stat cards used to come from
    1 - sum(holdings weights) -- whatever an analyst typed as target
    weights -- with zero connection to actual cash. HOLDINGS here sums to
    15% weight, so the old plug reported 85% cash; the real cashflow +
    trade ledger below says cash is actually ~5.6%. Once cashflow data
    exists, the dashboard's headline Equity/Cash must match the NAV
    panel's real rupee split, not the typed-weight plug."""
    up(client, "holdings", xlsx("Portfolio", H_HDR, HOLDINGS))   # weights sum to 15%
    up(client, "trades", xlsx("Trades", T_HDR, TRADES))
    up(client, "cashflows", xlsx("Cashflows", CF_HDR,
       [[dt.date(2026, 1, 1), "Contribution", 500000, "Seed capital"]]))

    nav = client.get("/api/nav").json()
    d = client.get("/api/dashboard").json()
    assert d["equityWeight"] == pytest.approx(nav["holdingsValue"] / nav["nav"])
    assert d["cashWeight"] == pytest.approx(nav["cashBalance"] / nav["nav"])
    assert d["equityWeight"] != pytest.approx(0.15, abs=0.01)   # not the stale weight plug


def test_sheets_row_building_handles_missing_upside_and_xirr(client):
    """No network here -- current_rows()/history_row() are pure functions
    over the same `derived` shape recalculate() already produces, checked
    against a book with no cashflows (xirr/upside can be None) so the
    row-building can't just assume every field is a number."""
    import sheets
    up(client, "holdings", xlsx("Portfolio", H_HDR, HOLDINGS))
    up(client, "trades", xlsx("Trades", T_HDR, TRADES))
    d = client.get("/api/dashboard").json()
    nav = client.get("/api/nav").json()
    pos = client.get("/api/positions").json()
    derived = {"generatedAt": "2026-01-01T00:00:00+00:00", "dashboard": d, "nav": nav, "positions": pos}

    rows = sheets.current_rows(derived)
    assert rows[0][0] == "DL India Core -- current snapshot"
    header_row = next(r for r in rows if r and r[0] == "Ticker")
    tickers = {r[0] for r in rows[rows.index(header_row) + 1:]}
    assert tickers == {"ALPHACHEM", "BETAFIN"}

    row = sheets.history_row(derived)
    assert row[0] == dt.date.today().isoformat()
    assert row[1] == pytest.approx(nav["nav"])


def test_sheets_ledger_rows_mirror_the_raw_book_no_computed_columns(client):
    """holdings_rows()/trades_rows()/cashflows_rows() must be the book as
    uploaded -- no CMP, no P&L, no status -- since these tabs exist
    specifically so raw and computed data aren't mixed in one place."""
    import sheets
    book = {
        "holdings": [{"tk": "ALPHACHEM", "co": "Alpha Chemicals", "sector": "Materials",
                      "industry": "Specialty", "analyst": "Dev", "qty": 400, "wt": 0.08,
                      "fullWt": 0.10, "tp": 620, "addLvl": 480, "conv": "High",
                      "thesis": "t", "strategy": "s", "yfSymbol": None}],
        "trades": [{"date": "2026-01-15", "tk": "ALPHACHEM", "side": "Buy", "qty": 400,
                    "price": 505.0, "costs": 202.0}],
        "cashflows": [{"date": "2026-01-01", "type": "Contribution", "amount": 500000, "note": "Seed"}],
    }
    h_rows = sheets.holdings_rows(book)
    assert h_rows[0] == sheets.HOLDINGS_HEADER
    assert h_rows[1][0] == "ALPHACHEM" and h_rows[1][6] == 8.0   # weight written as percent, not fraction
    assert "CMP" not in sheets.HOLDINGS_HEADER and "Status" not in sheets.HOLDINGS_HEADER

    t_rows = sheets.trades_rows(book)
    assert t_rows == [sheets.TRADES_HEADER, ["2026-01-15", "ALPHACHEM", "Buy", 400, 505.0, 202.0]]

    c_rows = sheets.cashflows_rows(book)
    assert c_rows == [sheets.CASHFLOWS_HEADER, ["2026-01-01", "Contribution", 500000, "Seed"]]


def test_export_history_503s_when_the_sheet_is_unreachable(client):
    """Under the local test backend there's no Google Sheets API at all --
    read_all_tabs() returns None, and the endpoint must say so with a
    503, not a 200 full of nothing or an unhandled crash."""
    r = client.get("/api/export/history")
    assert r.status_code == 503


def test_export_history_returns_every_tab_when_the_sheet_is_reachable(client, monkeypatch):
    import sheets
    fake = {
        sheets.TAB: [sheets.HISTORY_HEADER, ["2026-01-01", 100000, 5000, 95000, 0, 0, 0, 5.0, 95.0, 0.1, 3]],
        "Holdings": sheets.holdings_rows({"holdings": []}),
        "Trades": sheets.trades_rows({"holdings": [], "trades": []}),
        "Cashflow": sheets.cashflows_rows({"holdings": [], "cashflows": []}),
    }
    monkeypatch.setattr(sheets, "read_all_tabs", lambda name: fake if name == sheets.HISTORY_SHEET else None)

    r = client.get("/api/export/history")
    assert r.status_code == 200
    wb = openpyxl.load_workbook(io.BytesIO(r.content))
    assert set(wb.sheetnames) == {sheets.TAB, "Holdings", "Trades", "Cashflow"}
    assert [c.value for c in next(wb[sheets.TAB].iter_rows())] == sheets.HISTORY_HEADER


def test_cashflows_export_round_trips(client):
    up(client, "holdings", xlsx("Portfolio", H_HDR, HOLDINGS))
    up(client, "trades", xlsx("Trades", T_HDR, TRADES))
    up(client, "cashflows", xlsx("Cashflows", CF_HDR,
       [[dt.date(2026, 1, 1), "Contribution", 500000, "Seed capital"]]))
    before = client.get("/api/nav").json()

    cf_xlsx = client.get("/api/export/cashflows").content
    r = up(client, "cashflows", cf_xlsx)
    assert r.json()["added"] == 0 and r.json()["duplicates"] == 1   # same row, re-imported

    after = client.get("/api/nav").json()
    assert after == before


def test_holdings_reimport_does_not_touch_trade_history(client):
    up(client, "holdings", xlsx("Portfolio", H_HDR, HOLDINGS))
    up(client, "trades", xlsx("Trades", T_HDR, TRADES))
    up(client, "holdings", xlsx("Portfolio", H_HDR, HOLDINGS))
    assert client.get("/api/performance").json()["tradeCount"] == 2


# --------------------------------------------------------------- maths
def test_pl_matches_a_hand_calculation(client):
    up(client, "holdings", xlsx("Portfolio", H_HDR, HOLDINGS))
    up(client, "trades", xlsx("Trades", T_HDR, TRADES))
    d = client.get("/api/dashboard").json()

    invested = 400 * 505 + 900 * 298
    value = 400 * 542 + 900 * 305
    costs = 202.0 + 268.2
    assert d["invested"] == pytest.approx(invested)
    assert d["currentValue"] == pytest.approx(value)
    assert d["unrealized"] == pytest.approx(value - invested)
    assert d["realized"] == 0
    assert d["overallPL"] == pytest.approx(value - invested - costs)
    # the identity that must always hold
    assert d["overallPL"] == pytest.approx(d["unrealized"] + d["realized"] - d["costs"])


def test_a_later_buy_never_rewrites_an_earlier_sales_realized_pl(client):
    """The exact bug the CIO review caught: buy 100 @ 10, sell 50 @ 15
    (books Rs 250), then buy 100 more @ 20. Under a lifetime-average-cost
    scheme the later buy drags the average up and silently zeroes out the
    profit already booked on the sale. FIFO must not let that happen --
    the Rs 250 stays Rs 250 regardless of what's bought afterward. The
    remaining position's avg cost should reflect only the lots still
    held (50 @ 10 + 100 @ 20), not a lifetime blend."""
    up(client, "holdings", xlsx("Portfolio", H_HDR, HOLDINGS))
    up(client, "trades", xlsx("Trades", T_HDR, [
        [dt.date(2026, 1, 1), "ALPHACHEM", "Buy",  100, 10.0, 0.0],
        [dt.date(2026, 1, 5), "ALPHACHEM", "Sell", 50,  15.0, 0.0],
        [dt.date(2026, 1, 10), "ALPHACHEM", "Buy", 100, 20.0, 0.0],
    ]))
    row = next(r for r in client.get("/api/performance").json()["positions"] if r["tk"] == "ALPHACHEM")
    assert row["realized"] == pytest.approx(250)
    assert row["net"] == pytest.approx(150)
    assert row["avg"] == pytest.approx((50 * 10 + 100 * 20) / 150)


def test_holdings_qty_drives_unrealized_pl_and_flags_mismatch(client):
    """The holdings file's Qty column, not the trades ledger, now sizes
    unrealized P&L -- realized P&L and avg cost stay trades-only. A stated
    qty that disagrees with what the trades imply must be reported, not
    silently overridden."""
    hdr = H_HDR[:5] + ["Qty"] + H_HDR[5:]
    rows = [HOLDINGS[0][:5] + [350] + HOLDINGS[0][5:],   # trades say 400 bought -> mismatch
            HOLDINGS[1][:5] + [900] + HOLDINGS[1][5:]]   # trades say 900 bought -> matches
    up(client, "holdings", xlsx("Portfolio", hdr, rows))
    up(client, "trades", xlsx("Trades", T_HDR, TRADES))
    d = client.get("/api/dashboard").json()

    unreal = 350 * (542.0 - 505.0) + 900 * (305.0 - 298.0)
    assert d["unrealized"] == pytest.approx(unreal)
    assert d["realized"] == 0                             # unaffected, still trades-only
    assert d["qtyMismatches"] == [{"tk": "ALPHACHEM", "holdingsQty": 350, "tradesQty": 400}]


def test_repeated_holdings_and_trades_imports_stay_consistent(client):
    """Real usage is never one clean import -- it's holdings replaced a few
    times, trades files re-uploaded (with overlap), a ticker dropped from
    holdings while its trade history stays on record, and a brand new
    ticker showing up only in a trades file. None of that should corrupt
    the book or leave stale numbers behind."""
    qh = H_HDR[:5] + ["Qty"] + H_HDR[5:]

    # Round 1: two names, holdings qty == trades qty, no mismatch.
    up(client, "holdings", xlsx("Portfolio", qh,
        [HOLDINGS[0][:5] + [400] + HOLDINGS[0][5:],    # ALPHACHEM qty 400
         HOLDINGS[1][:5] + [900] + HOLDINGS[1][5:]]))  # BETAFIN qty 900
    up(client, "trades", xlsx("Trades", T_HDR, TRADES))
    assert client.get("/api/dashboard").json()["qtyMismatches"] == []

    # Round 2: holdings file REPLACED -- BETAFIN dropped, GAMMASTL added,
    # ALPHACHEM's qty trimmed to 350 (a sell the trades file hasn't caught
    # up with yet -- deliberate transient mismatch).
    up(client, "holdings", xlsx("Portfolio", qh, [
        ["ALPHACHEM", "Alpha Chemicals", "Materials", "Specialty", "Dev", 350, 5.0, 10.0, 650, 480, "High", "t", "s"],
        ["GAMMASTL", "Gamma Steel", "Materials", "Long", "Dev", 200, 4.0, 6.0, 150, None, "Medium", "t", "s"],
    ]))
    mid = client.get("/api/dashboard").json()
    assert mid["qtyMismatches"] == [{"tk": "ALPHACHEM", "holdingsQty": 350, "tradesQty": 400}]

    # The trades file that resolves it: sell the 50 ALPHACHEM, buy GAMMASTL.
    trades_b = xlsx("Trades", T_HDR, [
        [dt.date(2026, 3, 1), "ALPHACHEM", "Sell", 50, 600.0, 25.0],
        [dt.date(2026, 3, 2), "GAMMASTL", "Buy", 200, 140.0, 0.0],
    ])
    r = up(client, "trades", trades_b)
    assert r.json()["added"] == 2
    assert client.get("/api/dashboard").json()["qtyMismatches"] == []   # resolved

    # Round 3: re-upload the same trades file and the same holdings file --
    # pure duplicates / a no-op replace, must not double-count anything.
    r = up(client, "trades", trades_b)
    assert r.json()["added"] == 0 and r.json()["duplicates"] == 2 and r.json()["total"] == 4
    up(client, "holdings", xlsx("Portfolio", qh, [
        ["ALPHACHEM", "Alpha Chemicals", "Materials", "Specialty", "Dev", 350, 5.0, 10.0, 650, 480, "High", "t", "s"],
        ["GAMMASTL", "Gamma Steel", "Materials", "Long", "Dev", 200, 4.0, 6.0, 150, None, "Medium", "t", "s"],
    ]))
    book = drive.read_json(drive.PORTFOLIO_JSON)
    assert len(book["holdings"]) == 2 and len(book["trades"]) == 4

    # Round 4: a trade for a ticker in neither holdings file -> placeholder.
    r = up(client, "trades", xlsx("Trades", T_HDR,
        [[dt.date(2026, 3, 5), "DELTAENG", "Buy", 50, 100.0, 0.0]]))
    assert r.json()["placeholderHoldings"] == ["DELTAENG"]

    d = client.get("/api/dashboard").json()
    unreal = 350 * (542.0 - 505.0) + 900 * (305.0 - 298.0) + 200 * 0 + 50 * 0
    realized = 50 * 600.0 - 50 * 505.0
    costs = (202.0 + 25.0) + 268.2 + 0.0 + 0.0
    assert d["unrealized"] == pytest.approx(unreal)
    assert d["realized"] == pytest.approx(realized)
    assert d["costs"] == pytest.approx(costs)
    assert d["overallPL"] == pytest.approx(d["unrealized"] + d["realized"] - d["costs"])
    assert d["qtyMismatches"] == []

    # BETAFIN was dropped from the holdings file two rounds ago: its P&L
    # still counts (its trade is still on the ledger) but it must not show
    # up as a position any more.
    tickers = {p["tk"] for p in client.get("/api/positions").json()["positions"]}
    assert tickers == {"ALPHACHEM", "GAMMASTL", "DELTAENG"}


def test_exiting_status_beats_a_target_hit(client):
    up(client, "holdings", xlsx("Portfolio", H_HDR, HOLDINGS))
    pos = {p["tk"]: p for p in client.get("/api/positions").json()["positions"]}
    assert pos["BETAFIN"]["status"]["text"] == "EXITING"      # fullWt == 0 wins
    assert pos["ALPHACHEM"]["status"]["text"] == "ON THESIS"


def test_risk_has_four_scenarios_and_excludes_unpriced(client):
    up(client, "holdings", xlsx("Portfolio", H_HDR, HOLDINGS))
    r = client.get("/api/risk").json()
    assert [s["key"] for s in r["scenarios"]] == ["mid", "give1m", "low", "worst"]
    assert r["priced"] == 2 and r["weightedBeta"] is not None


def test_liquidity_view_flags_days_to_liquidate_and_tier(client):
    """ALPHACHEM: 400 shares held against an ADV of 500 -> 80% of a day's
    volume, which at a 20% participation rate is 4 trading days to exit --
    'Tight'. BETAFIN: 900 against an ADV of 5000 -> 18%, 0.9 days --
    'Liquid'. Real volume data, not a made-up slippage percentage."""
    up(client, "holdings", xlsx("Portfolio", H_HDR, HOLDINGS))
    up(client, "trades", xlsx("Trades", T_HDR, TRADES))
    liq = {r["tk"]: r for r in client.get("/api/risk").json()["liquidity"]["rows"]}

    a = liq["ALPHACHEM"]
    assert a["positionOverAdv"] == pytest.approx(0.8)
    assert a["daysToLiquidate"] == pytest.approx(4.0)
    assert a["tier"] == "Tight"

    b = liq["BETAFIN"]
    assert b["daysToLiquidate"] == pytest.approx(0.9)
    assert b["tier"] == "Liquid"


def test_give1m_survives_a_stock_that_went_to_zero():
    g = calculator.SCENARIOS[1][3]
    assert g({"r1m": -1.0}) is None          # 1 + r1m == 0 would divide by zero
    assert g({"r1m": 0.05}) == pytest.approx(1 / 1.05 - 1)


def test_xirr_matches_hand_calculation():
    """Rs 1,00,000 invested, Rs 1,10,000 back exactly one year later ->
    XIRR must be exactly 10%, no numerical wobble. Two years compounding
    at the same 10% must give back Rs 1,21,000 -- and solve back to 10%
    too, proving the solver handles more than a single-period trivial case."""
    r = calculator.xirr([(dt.date(2025, 1, 1), -100000), (dt.date(2026, 1, 1), 110000)])
    assert r == pytest.approx(0.10, abs=1e-4)

    r2 = calculator.xirr([(dt.date(2025, 1, 1), -100000), (dt.date(2027, 1, 1), 121000)])
    assert r2 == pytest.approx(0.10, abs=1e-4)

    # multiple contributions -- no closed form, so verify NPV(r) actually
    # lands at zero at the solved rate rather than trusting a magic number
    flows = [(dt.date(2024, 1, 1), -50000), (dt.date(2025, 1, 1), -50000),
             (dt.date(2026, 1, 1), 115000)]
    r3 = calculator.xirr(flows)
    t0 = min(d for d, _ in flows)
    npv = sum(a / (1 + r3) ** ((d - t0).days / 365.0) for d, a in flows)
    assert abs(npv) < 1.0, f"NPV at solved rate should be ~0, got {npv}"

    assert calculator.xirr([]) is None
    assert calculator.xirr([(dt.date(2025, 1, 1), -100000)]) is None, "one flow implies no rate"
    assert calculator.xirr([(dt.date(2025, 1, 1), -100000), (dt.date(2026, 1, 1), -5000)]) is None, \
        "money that only ever left implies no rate"


def test_benchmark_xirr_replays_the_same_flows_into_the_index():
    """Rs 1,00,000 contributed when the index was at 100 -> buys 1,000
    'units'. A year later the index is at 120 -> shadow value is exactly
    Rs 1,20,000, a clean 20% -- and that must match a hand-computed XIRR
    of a Rs 100,000-in/Rs 120,000-out pair exactly."""
    levels = {"2025-01-01": 100.0, "2026-01-01": 120.0}
    flows = [(dt.date(2025, 1, 1), -100000.0)]
    r = calculator.benchmark_xirr(flows, levels)
    expected = calculator.xirr([(dt.date(2025, 1, 1), -100000.0), (dt.date(2026, 1, 1), 120000.0)])
    assert r == pytest.approx(expected, abs=1e-6) == pytest.approx(0.20, abs=1e-4)

    assert calculator.benchmark_xirr([], levels) is None
    assert calculator.benchmark_xirr(flows, {}) is None
    # a contribution before the earliest level we have must not guess
    assert calculator.benchmark_xirr([(dt.date(2020, 1, 1), -1000.0)], levels) is None


def test_insights_never_mention_stocks_not_in_the_book():
    """The exact bug the CIO review caught: 'BEL + Kaynes' and 'Sun,
    Titan' used to be hardcoded into the Crowding and Posture insight
    text and fire on any book crossing their threshold -- a client could
    see analysis about stocks they don't own. Every name in the output
    must now come from the actual holdings passed in."""
    holdings = [
        {"tk": "STOCKA", "co": "Stock A", "sector": "Materials", "wt": 0.30,
         "fullWt": 0.40, "conv": "High", "addLvl": 100},
        {"tk": "STOCKB", "co": "Stock B", "sector": "Materials", "wt": 0.20,
         "fullWt": 0.20, "conv": "Medium", "addLvl": None},
    ]
    signals = {"STOCKA": {"val": -1.0, "momo": 0.0}, "STOCKB": {"val": -1.0, "momo": 0.0}}
    bench_sect = {"Materials": 0.05}
    P = {"rows": [{"tk": "STOCKA", "total": 100, "momo": 0.0},
                  {"tk": "STOCKB", "total": 50, "momo": 0.0}], "total": 150}
    C = {"cash": 0.5, "pendingAdds": 0.10, "exits": 0.0, "trims": 0.0, "fullBuild": 0.60,
         "convictionWeights": {"High": 0.30, "Medium": 0.20, "Low": 0.0},
         "equity": 0.50, "top3": ["STOCKA", "STOCKB"]}

    out = calculator.insights(holdings, signals, bench_sect, P, C)
    blob = " ".join(i["text"] for i in out)
    for fake in ("BEL", "Kaynes", "Sun,", "Titan"):
        assert fake not in blob, f"hardcoded mock name {fake!r} leaked into insights"

    crowding = next(i for i in out if i["tag"] == "Crowding")
    assert "Materials" in crowding["text"] and "STOCKA" in crowding["text"]
    posture = next(i for i in out if i["tag"] == "Posture")
    assert "STOCKA" in posture["text"]          # the actual staged name, not a mock one


def test_size_gap_is_reported_not_hidden(client):
    """A holding with a price but no market cap belongs to no size bucket,
    so the bars total less than 100%. Say so rather than leave a hole."""
    feed = json.loads(json.dumps(FEED))
    feed["signals"]["BETAFIN"]["mcap"] = None
    drive.write_json(drive.SIGNALS_JSON, feed)
    up(client, "holdings", xlsx("Portfolio", H_HDR, HOLDINGS))
    e = client.get("/api/exposure").json()
    assert e["sizeUnclassified"] > 0
    assert sum(s["portfolio"] for s in e["size"]) + e["sizeUnclassified"] == pytest.approx(1.0)


def test_exposure_benchmark_aliases_indian_sector_labels_to_yahoos(client):
    """Real bug this fixes: Yahoo/Screener.in report sectors as 'Basic
    Materials', an analyst types 'Materials' in the holdings file. Without
    aliasing, that mismatch silently priced the benchmark at 0% for the
    sector and listed it under 'not held' even though the book plainly
    holds it."""
    feed = json.loads(json.dumps(FEED))
    feed["bench_sect"] = {"Basic Materials": 0.12, "Financial Services": 0.30}
    drive.write_json(drive.SIGNALS_JSON, feed)
    up(client, "holdings", xlsx("Portfolio", H_HDR, HOLDINGS))   # ALPHACHEM: Materials, BETAFIN: Financials

    e = client.get("/api/exposure").json()
    by_sector = {s["sector"]: s for s in e["sectors"]}
    assert by_sector["Materials"]["benchmark"] == pytest.approx(0.12)
    assert by_sector["Financials"]["benchmark"] == pytest.approx(0.30)
    assert e["notHeld"]["count"] == 0    # both benchmark sectors matched via alias, neither is "not held"


# ------------------------------------------------------- error handling
def test_missing_sheet_gives_a_useful_422(client):
    r = up(client, "trades", xlsx("Wrong", T_HDR, TRADES))
    assert r.status_code == 422 and "No 'Trades' sheet" in r.json()["detail"]


def test_validate_endpoint_matches_import_errors_but_writes_nothing(client):
    """The real bug this exists to prevent: holdings + trades upload
    together, holdings parses fine and trades doesn't -- without
    validating first, holdings would already be saved by the time trades
    fails, leaving a red error on screen with no hint that half the
    upload actually went through. /api/validate/* must reject the exact
    same bad file /api/import/* would, but touch nothing."""
    bad_trades = xlsx("Wrong", T_HDR, TRADES)
    r = client.post("/api/validate/trades", files={"file": ("f.xlsx", bad_trades, "application/octet-stream")})
    assert r.status_code == 422 and "No 'Trades' sheet" in r.json()["detail"]
    assert client.get("/api/dashboard").json()["tradesCount"] == 0   # nothing written

    good_holdings = xlsx("Portfolio", H_HDR, HOLDINGS)
    r2 = client.post("/api/validate/holdings", files={"file": ("f.xlsx", good_holdings, "application/octet-stream")})
    assert r2.status_code == 200 and r2.json()["errors"] == []
    assert client.get("/api/dashboard").json()["holdingsCount"] == 0   # still nothing written -- validate only


def test_missing_required_column_gives_a_useful_422(client):
    r = up(client, "trades", xlsx("Trades", ["Date", "Ticker", "Side"], [[None, "A", "Buy"]]))
    assert r.status_code == 422 and "Ticker, Side, Qty and Price" in r.json()["detail"]


def test_holdings_without_target_column_rejected(client):
    r = up(client, "holdings", xlsx("Portfolio", ["Ticker", "Company", "Weight"], [["A", "Alpha", 5]]))
    assert r.status_code == 422 and "Target column" in r.json()["detail"]


def test_sector_read_as_ticker_is_caught(client):
    r = up(client, "holdings", xlsx("Portfolio", ["Ticker", "Company", "Weight", "Target"],
                                     [["Specialty Chemicals", "Alpha", 5, 600]]))
    assert r.status_code == 422        # every row rejected -> no valid rows


def test_bad_row_reports_its_number_without_aborting_the_batch(client):
    body = up(client, "trades", xlsx("Trades", T_HDR, [
        [dt.date(2026, 5, 1), "AAA", "Buy", -5, 500, 0],
        [dt.date(2026, 5, 2), "AAA", "Buy", 5, 500, 0]])).json()
    assert body["added"] == 1
    assert body["errors"][0]["row"] == 1 and "> 0" in body["errors"][0]["message"]


def test_empty_and_garbage_uploads_rejected(client):
    assert up(client, "trades", b"").status_code == 422
    assert up(client, "holdings", b"not a workbook").status_code == 422


def test_a_failed_import_does_not_corrupt_what_is_stored(client):
    up(client, "holdings", xlsx("Portfolio", H_HDR, HOLDINGS))
    up(client, "trades", xlsx("Trades", T_HDR, TRADES))
    before = json.dumps(drive.read_json(drive.PORTFOLIO_JSON), sort_keys=True)

    assert up(client, "holdings", b"garbage").status_code == 422
    assert up(client, "trades", xlsx("Wrong", T_HDR, TRADES)).status_code == 422

    assert json.dumps(drive.read_json(drive.PORTFOLIO_JSON), sort_keys=True) == before
    assert client.get("/api/dashboard").json()["holdingsCount"] == 2


def test_drive_failure_surfaces_as_503_not_a_crash(client, monkeypatch):
    monkeypatch.setattr(drive.store(), "read_json",
                        lambda name: (_ for _ in ()).throw(drive.DriveError("Drive is down")))
    r = client.get("/api/dashboard")
    assert r.status_code == 503 and "Drive is down" in r.json()["detail"]


def test_store_reconnects_lazily_if_startup_never_ran(client, monkeypatch):
    """The real bug this fixes: some serverless ASGI runtimes never fire
    FastAPI's startup/lifespan event at all, so drive.connect() never runs
    and every request used to fail with a generic 'not connected' error
    forever -- not the actual reason (missing creds, API disabled, ...),
    and not self-healing. store() must reconnect on first use instead of
    just trusting startup already happened."""
    drive._store = None   # simulate startup never having fired
    r = client.get("/api/dashboard")
    assert r.status_code == 200
    assert drive._store is not None   # reconnected, not left broken


def test_archival_copy_failure_does_not_mask_a_successful_import(client, monkeypatch):
    """The exact bug found while building the cashflows feature: the
    archival .xlsx write is best-effort (nothing reads it to compute
    anything) and runs concurrently with the real portfolio.json write.
    If ONLY the archival write fails, the request must still succeed --
    the critical write already landed, and reporting failure for a book
    that was actually saved is worse than the archival copy being stale.

    write_json is implemented on top of write_bytes, so the failure must
    be scoped to just the archival filename -- a blanket patch would also
    break the critical portfolio.json write, which is a different bug."""
    real_write_bytes = drive.store().write_bytes
    def flaky_write_bytes(name, data, mime=None):
        if name == drive.HOLDINGS_XLSX:
            raise drive.DriveError("no quota")
        return real_write_bytes(name, data, mime)
    monkeypatch.setattr(drive.store(), "write_bytes", flaky_write_bytes)

    r = up(client, "holdings", xlsx("Portfolio", H_HDR, HOLDINGS))
    assert r.status_code == 200, r.text
    assert r.json()["holdings"] == 2
    assert client.get("/api/dashboard").json()["holdingsCount"] == 2   # the real write landed


def test_signals_refresh_trigger_fires_only_when_configured(client, monkeypatch):
    """No GITHUB_TOKEN/GITHUB_REPO set (the default in this test env) ->
    silent no-op, same as the other optional integrations. Set both ->
    a real dispatch call fires against the GitHub API on import."""
    import api
    calls = []
    monkeypatch.setattr(api.urllib.request, "urlopen", lambda req, timeout=10: calls.append(req))

    r = up(client, "holdings", xlsx("Portfolio", H_HDR, HOLDINGS))
    assert calls == []   # not configured -- must not attempt a call
    # ...and must SAY it didn't, naming what's missing: a silently skipped
    # refresh is indistinguishable from a working one, which is exactly
    # how stale factor scores went unnoticed on a real deployment.
    assert "not configured" in r.json()["signalRefresh"]
    assert "GITHUB_TOKEN" in r.json()["signalRefresh"]

    # Vercel publishes the repo/branch it deployed, so a token alone is
    # enough -- GITHUB_REPO going unset (while the token was fine) is the
    # exact failure seen on the real deployment.
    monkeypatch.setenv("GITHUB_TOKEN", "t")
    monkeypatch.setenv("VERCEL_GIT_REPO_OWNER", "jprasham")
    monkeypatch.setenv("VERCEL_GIT_REPO_SLUG", "dl-india-portfolio-dashboard")
    monkeypatch.setenv("VERCEL_GIT_COMMIT_REF", "main")
    r = up(client, "holdings", xlsx("Portfolio", H_HDR, HOLDINGS))
    assert r.json()["signalRefresh"] == "queued"
    assert calls[-1].full_url == ("https://api.github.com/repos/jprasham/dl-india-portfolio-dashboard"
                                   "/actions/workflows/update-signals.yml/dispatches")
    assert json.loads(calls[-1].data)["ref"] == "main"
    for v in ("VERCEL_GIT_REPO_OWNER", "VERCEL_GIT_REPO_SLUG", "VERCEL_GIT_COMMIT_REF"):
        monkeypatch.delenv(v)
    calls.clear()

    # Deliberately messy values -- a quoted, padded repo URL is what
    # actually ends up pasted into a hosting dashboard's env-var box, and
    # it used to 404 in a way indistinguishable from a bad token.
    monkeypatch.setenv("GITHUB_TOKEN", "  t\n"), monkeypatch.setenv("GITHUB_REPO", ' "https://github.com/me/repo/" ')
    r = up(client, "holdings", xlsx("Portfolio", H_HDR, HOLDINGS))
    assert len(calls) == 1
    assert calls[0].full_url == "https://api.github.com/repos/me/repo/actions/workflows/update-signals.yml/dispatches"
    assert r.json()["signalRefresh"] == "queued"
    # urllib defaults a POST body to x-www-form-urlencoded; the GitHub API
    # needs JSON here, and the mismatch was silent.
    assert calls[0].get_header("Content-type") == "application/json"
    assert calls[0].get_header("Authorization") == "Bearer t"       # whitespace stripped
    assert json.loads(calls[0].data)["ref"] == "main"

    # a rejected dispatch (bad token, wrong ref, ...) must surface the
    # reason and the repo/ref it tried, not vanish into a server log
    def boom(req, timeout=10):
        raise urllib.error.HTTPError(req.full_url, 404, "Not Found", {}, None)
    monkeypatch.setattr(api.urllib.request, "urlopen", boom)
    r = up(client, "holdings", xlsx("Portfolio", H_HDR, HOLDINGS))
    status = r.json()["signalRefresh"]
    assert status.startswith("failed:") and "404" in status and "me/repo" in status


def test_quick_price_refresh_patches_cmp_but_preserves_factor_scores(client, monkeypatch):
    """The real point of quick_prices(): update cmp/mcap/lo/hi for the
    book's own holdings immediately, WITHOUT touching val/momo/qual/beta
    -- those need the full 750-symbol cross-sectional job to mean
    anything and must survive untouched until that job next runs."""
    import build_signals
    monkeypatch.setattr(build_signals, "quick_prices",
                        lambda holdings: ({"ALPHACHEM": {"cmp": 999.0, "mcap": 5000, "lo": 400.0, "hi": 1000.0}}, []))

    up(client, "holdings", xlsx("Portfolio", H_HDR, HOLDINGS))

    sig = client.get("/api/signals").json()["signals"]["ALPHACHEM"]
    assert sig["cmp"] == 999.0 and sig["mcap"] == 5000                 # patched
    assert sig["val"] == FEED["signals"]["ALPHACHEM"]["val"]           # untouched
    assert sig["beta"] == FEED["signals"]["ALPHACHEM"]["beta"]         # untouched
    # the P&L on the very same response already reflects the new price
    pos = next(p for p in client.get("/api/positions").json()["positions"] if p["tk"] == "ALPHACHEM")
    assert pos["cmp"] == 999.0


def test_export_trades_filters_by_date_range(client):
    up(client, "holdings", xlsx("Portfolio", H_HDR, HOLDINGS))
    up(client, "trades", xlsx("Trades", T_HDR, TRADES))   # Jan 15 and Feb 3

    only_jan = client.get("/api/export/trades", params={"to_date": "2026-01-31"}).content
    rows = [list(r) for r in openpyxl.load_workbook(io.BytesIO(only_jan)).active.iter_rows(values_only=True)]
    assert [r[1] for r in rows[1:]] == ["ALPHACHEM"]   # only the Jan 15 row

    none = client.get("/api/export/trades", params={"from_date": "2026-06-01"}).content
    rows = [list(r) for r in openpyxl.load_workbook(io.BytesIO(none)).active.iter_rows(values_only=True)]
    assert len(rows) == 1   # header only

    both = client.get("/api/export/trades").content
    rows = [list(r) for r in openpyxl.load_workbook(io.BytesIO(both)).active.iter_rows(values_only=True)]
    assert len(rows) == 3   # header + both trades


def test_export_cashflows_filters_by_date_range(client):
    up(client, "holdings", xlsx("Portfolio", H_HDR, HOLDINGS))
    up(client, "cashflows", xlsx("Cashflows", CF_HDR, [
        [dt.date(2026, 1, 1), "Contribution", 500000, "Seed"],
        [dt.date(2026, 3, 1), "Withdrawal", 20000, "Pull"],
    ]))

    only_q1_jan = client.get("/api/export/cashflows",
                             params={"from_date": "2026-01-01", "to_date": "2026-01-31"}).content
    rows = [list(r) for r in openpyxl.load_workbook(io.BytesIO(only_q1_jan)).active.iter_rows(values_only=True)]
    assert [r[1] for r in rows[1:]] == ["Contribution"]
