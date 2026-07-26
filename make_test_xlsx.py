"""Build Holdings_File_TEST.xlsx: the real book with its gaps filled in.

Holdings_File.xlsx is complete enough to render but leaves parts of the
dashboard dark. This writes a copy with every blank populated, plus the Trades
sheet the real file doesn't have at all — which is the only reason Total P&L,
Realized, Unrealized and all four attribution panels currently read zero.

Real analyst content is copied verbatim. Invented text is tagged [placeholder]
so nobody mistakes it for someone's actual view.

    python make_test_xlsx.py
"""
import json
from datetime import date, timedelta

import openpyxl

SRC, OUT = "Holdings_File.xlsx", "Holdings_File_TEST.xlsx"
NAV = 10_000_000          # ₹1 crore book, so a 2.5% position is ₹2.5L
ADD_PULLBACK = 0.92       # stage adds 8% below spot
COST_RATE = 0.001         # 10 bps all-in

# Gaps in the real sheet. Target/AddLevel are derived from live prices below.
FILL = {
    "AIMTRON": {"Conviction": "High",
                "Strategy": "[placeholder] Technical setup: EMS order-book visibility. "
                            "Action: accumulate on pullbacks to the 50-day EMA."},
    "J&KBANK": {"Conviction": "Medium (Watchlist)",
                "Thesis": "[placeholder] Theme: J&K credit normalisation. Growth: loan book "
                          "recovery with improving asset quality. Valuation: trades below book, "
                          "re-rates as slippage falls.",
                "Strategy": "[placeholder] Technical setup: basing above long-term support. "
                            "Action: wait for a breakout on volume before sizing up."},
}
# Names to take partial profits on, so Realized P&L isn't zero either.
SELLS = {"MUTHOOTMF": 0.4, "IOLCP": 0.5, "SKIPPER": 0.35}


def entry_price(cmp_, lo):
    """A plausible historic fill: 35% of the way up from the 52-week low."""
    return round(lo + 0.35 * (cmp_ - lo), 2)


def build():
    wb = openpyxl.load_workbook(SRC)
    ws = wb["Portfolio"]
    sigs = json.load(open("signals.json", encoding="utf-8"))["signals"]
    hdr = [c.value for c in ws[1]]
    col = {name: i + 1 for i, name in enumerate(hdr)}

    filled, holdings = [], []
    for r in range(2, ws.max_row + 1):
        tk = ws.cell(r, col["Ticker"]).value
        if not tk:
            continue
        s = sigs.get(tk) or {}
        cmp_, lo = s.get("cmp"), s.get("lo")

        for field, value in FILL.get(tk, {}).items():
            if not ws.cell(r, col[field]).value:
                ws.cell(r, col[field]).value = value
                filled.append(f"{tk}.{field}")

        # Target: only PARKHOSPS lacks one. 25% above spot keeps it ON THESIS.
        if not ws.cell(r, col["Target"]).value and cmp_:
            ws.cell(r, col["Target"]).value = round(cmp_ * 1.25, 2)
            filled.append(f"{tk}.Target")
        # AddLevel is blank for every row, which reads as "event-gated" on the
        # dashboard. Filling it turns the staged-add triggers into real prices.
        if not ws.cell(r, col["AddLevel"]).value and cmp_:
            ws.cell(r, col["AddLevel"]).value = round(cmp_ * ADD_PULLBACK, 2)
            filled.append(f"{tk}.AddLevel")

        wt = ws.cell(r, col["Weight"]).value
        if cmp_ and lo:
            holdings.append((tk, float(wt), cmp_, lo))

    # ---- Trades sheet: the real workbook has none, so P&L has nothing to compute
    tr = wb.create_sheet("Trades")
    tr.append(["Date", "Ticker", "Side", "Qty", "Price", "Costs"])
    today = date.today()
    for i, (tk, wt, cmp_, lo) in enumerate(holdings):
        px = entry_price(cmp_, lo)
        qty = max(1, round(NAV * wt / 100 / px))
        tr.append([today - timedelta(days=240 - i * 12), tk, "Buy", qty, px,
                   round(qty * px * COST_RATE, 2)])
    for i, (tk, frac) in enumerate(SELLS.items()):
        row = next(h for h in holdings if h[0] == tk)
        px = entry_price(row[2], row[3])
        qty = max(1, int(round(NAV * row[1] / 100 / px) * frac))
        sell_px = round(row[2] * 0.98, 2)
        tr.append([today - timedelta(days=45 - i * 15), tk, "Sell", qty, sell_px,
                   round(qty * sell_px * COST_RATE, 2)])
    for c, w in zip("ABCDEF", (12, 12, 8, 10, 12, 10)):
        tr.column_dimensions[c].width = w

    wb.save(OUT)
    return filled, len(holdings), tr.max_row - 1


if __name__ == "__main__":
    filled, n, trades = build()
    print(f"{OUT}: {n} holdings, {trades} trades, {len(filled)} blanks filled")
    for f in filled:
        print("  +", f)
